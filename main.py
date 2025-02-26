import requests
import time
import openpyxl
from openpyxl import Workbook
from datetime import datetime

workbook = Workbook()

def Getting_Game_Id():
    # Define the URL
    _url = "https://ca.1xbet.com/service-api/LiveFeed/GetSportsShortZip?sports=85&champs=2665392&lng=en&gr=828&country=85&virtualSports=true&groupChamps=true"

    # Send a GET request
    response = requests.get(_url)

    if response.status_code == 200:
        try:
            data = response.json()
            all_data = []

            # Loop through possible indices
            for idx in [20, 21, 22, 23, 24, 25, 26, 27]:
                try:
                    for i in range(5):
                        Game_Id = data["Value"][idx]["L"][2]["G"][i]["I"]
                        Score = data["Value"][idx]["L"][2]["G"][i]["SC"]["FS"]
                        Time_val = data["Value"][idx]["L"][2]["G"][i]["SC"]["TS"]

                        loop_data = {
                            "loop_id": f"loop{i}",
                            "Game_Id": Game_Id,
                            "Score": Score,
                            "Time": Time_val
                        }
                        all_data.append(loop_data)
                    break
                except (IndexError, KeyError):
                    continue

            empty_score_data = []

            for entry in all_data:
                if not entry["Score"]:
                    empty_score_data.append({
                        "loop_id": entry["loop_id"],
                        "Game_Id": entry["Game_Id"],
                        "Time": entry["Time"]
                    })

            if empty_score_data:
                lowest_time_entry = min(
                    empty_score_data,
                    key=lambda x: int(x["Time"]) if isinstance(x["Time"], (str, int)) else float("inf")
                )
                return lowest_time_entry['Game_Id']
            else:
                print("No entries with an empty Score were found.")
                return None

        except Exception as e:
            print(f"There was an error: {e}")
            return None
    else:
        print("Failed to get data from the server.")
        return None

def Recording():
    print("Started the Recording Phase")
    response = requests.get(url)
    ID_GAME = str(Getting_Game_Id())

    if response.status_code == 200:
        try:
            data = response.json()
            Team1 = data["Value"]["O1"]
            Team2 = data["Value"]["O2"]
            Time_val = int(data["Value"]["SC"]["TS"])
            number_of_iterations = 20

            # Construct sheet name with Game ID + Current Date
            current_date = datetime.today().strftime("%Y-%m-%d")
            new_sheet_name = f"{ID_GAME}_{current_date}"

            value_dict = {
                f'{Team1} Odd': {'W1_Odd': data["Value"]["GE"][0]["E"][0][0]["C"]},
                f'{Team2} Odd': {'W2_Odd': data["Value"]["GE"][0]["E"][2][0]["C"]},
                f'{Team1} Score': {'T1_Score': 0},
                f'{Team2} Score': {'T2_Score': 0},
                'X': {'X_Odd': data["Value"]["GE"][0]["E"][1][0]["C"]},
                'Time': {'Time': 0}  # First recorded time should be 0
            }

            print("Initial values:", value_dict)

            header = ['Iteration', f'{Team1} Score', f'{Team2} Score', f'{Team1} Odd', 'X', f'{Team2} Odd', 'Time']

            # Load or create workbook
            try:
                print("Opening Excel file...")
                wb = openpyxl.load_workbook('recorded_data.xlsx')
            except FileNotFoundError:
                wb = Workbook()

            if new_sheet_name in wb.sheetnames:
                sheet = wb[new_sheet_name]
            else:
                sheet = wb.create_sheet(title=new_sheet_name)

            # Append header row
            sheet.append(header)

            # Record the first row (Iteration 0) with Time set to 0
            initial_row = [0] + [''] * (len(header) - 1)
            for team, value in value_dict.items():
                if team in header:
                    idx = header.index(team)
                    inner_value = list(value.values())[0]
                    initial_row[idx] = inner_value
            sheet.append(initial_row)
            wb.save("recorded_data.xlsx")
            print("Initial data recorded.")

            # Start iterations from 1 onward
            for i in range(1, number_of_iterations):
                if i == 1:
                    sleep_time = Time_val + 60 if Time_val < 10 else Time_val + 40
                else:
                    sleep_time = 30

                print(f"Iteration {i}: Sleeping for {sleep_time} seconds...")
                time.sleep(sleep_time)

                response = requests.get(url)
                if response.status_code == 200:
                    data = response.json()
                    Time_val = int(data["Value"]["SC"]["TS"])
                    try:
                        s1_score = data["Value"]["SC"]["FS"]["S1"]
                    except (KeyError, IndexError):
                        s1_score = 0
                    try:
                        s2_score = data["Value"]["SC"]["FS"]["S2"]
                    except (KeyError, IndexError):
                        s2_score = 0

                    # Update value_dict
                    value_dict = {
                        f'{Team1} Odd': {'W1_Odd': data["Value"]["GE"][0]["E"][0][0]["C"]},
                        f'{Team2} Odd': {'W2_Odd': data["Value"]["GE"][0]["E"][2][0]["C"]},
                        f'{Team1} Score': {'T1_Score': s1_score},
                        f'{Team2} Score': {'T2_Score': s2_score},
                        'X': {'X_Odd': data["Value"]["GE"][0]["E"][1][0]["C"]},
                        'Time': {'Time': Time_val}
                    }
                    print("Updated values:", value_dict)

                    row = [i] + [''] * (len(header) - 1)
                    for team, value in value_dict.items():
                        if team in header:
                            idx = header.index(team)
                            inner_value = list(value.values())[0]
                            row[idx] = inner_value
                    sheet.append(row)
                    wb.save("recorded_data.xlsx")

                    if Time_val > 330:
                        print("Time is more than 30 seconds, breaking the loop.")
                        Record_Setup()
                        break
                else:
                    print("Failed to fetch data during iteration.")

        except Exception as e:
            print("There is an error somewhere:", str(e))
            Record_Setup()
    else:
        print("Failed to fetch data from the URL.")

def Record_Setup():
    print("Starting the Record Setup process")
    global url
    ID_GAME = Getting_Game_Id()
    current_date = datetime.today().strftime("%Y-%m-%d")
    url = f"https://ca.1xbet.com/service-api/LiveFeed/GetGameZip?id={ID_GAME}&lng=en&isSubGames=true&GroupEvents=true&countevents=250&grMode=4&topGroups=&country=85&marketType=1"
    print(url)
    response = requests.get(url)
    if response.status_code == 200:
        try:
            print("Getting the Time Value")
            data = response.json()
            Time_val = data["Value"]["SC"]["TS"]
            for i in range(Time_val):
                response = requests.get(url)
                if response.status_code == 200:
                    data = response.json()
                    Time_val = data["Value"]["SC"]["TS"]
                    print(Time_val)
                if Time_val < 32:
                    Recording()
                time.sleep(30)
        except Exception as e:
            print(f"There is an error somewhere: {e}")

Record_Setup()
