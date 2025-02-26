import threading
import requests
import time
import openpyxl
from openpyxl import Workbook

# Create a new workbook (or open an existing one)
workbook = Workbook()
# Global variable to hold the game URL (will be set in Record_Setup)
url = None

def Getting_Game_Id():
    # Define the URL to fetch game data
    _url = ("https://ca.1xbet.com/service-api/LiveFeed/GetSportsShortZip?"
            "sports=85&champs=2665392&lng=en&gr=828&country=85&virtualSports=true&groupChamps=true")
    response = requests.get(_url)
    if response.status_code == 200:
        try:
            data = response.json()
            all_data = []
            # Loop through possible indices to extract game data
            for idx in [20, 21, 22, 23, 24, 25, 26, 27]:
                try:
                    for i in range(5):
                        Game_Id = data["Value"][idx]["L"][2]["G"][i]["I"]
                        Score = data["Value"][idx]["L"][2]["G"][i]["SC"]["FS"]
                        Time_val = data["Value"][idx]["L"][2]["G"][i]["SC"]["TS"]
                        loop_data = {
                            "loop_id": f"loop{i}",
                            "Game_Id": Game_Id,
                            "Score": Score,
                            "Time": Time_val
                        }
                        all_data.append(loop_data)
                    break  # Exit the loop if data was successfully read
                except (IndexError, KeyError):
                    continue

            # Check for entries with an empty Score
            empty_score_data = [ 
                {"loop_id": entry["loop_id"], "Game_Id": entry["Game_Id"], "Time": entry["Time"]}
                for entry in all_data if not entry["Score"]
            ]
            if empty_score_data:
                lowest_time_entry = min(
                    empty_score_data,
                    key=lambda x: int(x["Time"]) if isinstance(x["Time"], (str, int)) else float("inf")
                )
                return lowest_time_entry['Game_Id']
            else:
                print("No entries with an empty Score were found.")
                return None

        except Exception as e:
            print(f"There was an error in Getting_Game_Id: {e}")
            return None
    else:
        print("Failed to get data from the server in Getting_Game_Id.")
        return None

def Recording():
    print("Started the Recording Phase")
    global url
    response = requests.get(url)
    ID_GAME = str(Getting_Game_Id())
    if response.status_code == 200:
        try:
            data = response.json()
            Team1 = data["Value"]["O1"]
            Team2 = data["Value"]["O2"]
            Time_val = int(data["Value"]["SC"]["TS"])
            number_of_iterations = 20

            # Prepare the initial dictionary of values
            value_dict = {
                f'{Team1} Odd': {'W1_Odd': data["Value"]["GE"][0]["E"][0][0]["C"]},
                f'{Team2} Odd': {'W2_Odd': data["Value"]["GE"][0]["E"][2][0]["C"]},
                f'{Team1} Score': {'T1_Score': 0},
                f'{Team2} Score': {'T2_Score': 0},
                'X': {'X_Odd': data["Value"]["GE"][0]["E"][1][0]["C"]},
                'Time': {'Time': Time_val}
            }
            print("Initial value dict:", value_dict)

            header = ['Iteration', f'{Team1} Score', f'{Team2} Score', f'{Team1} Odd', 'X', f'{Team2} Odd', 'Time']
            
            # Try to open an existing Excel file, otherwise create a new one
            try:
                print("Opening Excel file...")
                wb = openpyxl.load_workbook('recorded_data.xlsx')
                sheet = wb.active
            except FileNotFoundError:
                wb = Workbook()
                sheet = wb.active

            new_sheet_name = ID_GAME
            if new_sheet_name in wb.sheetnames:
                sheet = wb[new_sheet_name]
            else:
                sheet = wb.create_sheet(title=new_sheet_name)
            sheet.append(header)

            # Start recording data over multiple iterations
            for i in range(number_of_iterations):
                # For the first iteration, use the computed sleep time
                if i == 0:
                    sleep_time = Time_val + 55 if Time_val < 10 else Time_val + 35
                else:
                    sleep_time = 30

                print(f"Iteration {i+1}: Sleeping for {sleep_time} seconds...")
                time.sleep(sleep_time)

                response = requests.get(url)
                if response.status_code == 200:
                    data = response.json()
                    Time_val = int(data["Value"]["SC"]["TS"])
                    try:
                        s1_score = data["Value"]["SC"]["FS"]["S1"]
                    except (KeyError, IndexError):
                        s1_score = 0
                    try:
                        s2_score = data["Value"]["SC"]["FS"]["S2"]
                    except (KeyError, IndexError):
                        s2_score = 0

                    # Update dictionary with new values
                    value_dict = {
                        f'{Team1} Odd': {'W1_Odd': data["Value"]["GE"][0]["E"][0][0]["C"]},
                        f'{Team2} Odd': {'W2_Odd': data["Value"]["GE"][0]["E"][2][0]["C"]},
                        f'{Team1} Score': {'T1_Score': s1_score},
                        f'{Team2} Score': {'T2_Score': s2_score},
                        'X': {'X_Odd': data["Value"]["GE"][0]["E"][1][0]["C"]},
                        'Time': {'Time': Time_val}
                    }
                    print("Updated value dict:", value_dict)
                    # Prepare the row with the first column as iteration index
                    row = [i] + [''] * (len(header) - 1)
                    for team, value in value_dict.items():
                        if team in header:
                            idx = header.index(team)
                            inner_value = list(value.values())[0]
                            row[idx] = inner_value
                    sheet.append(row)
                    wb.save("recorded_data.xlsx")

                    # Optionally, break the loop if a condition is met
                    if Time_val > 330:
                        print("Time is more than 30 seconds. Stopping recording.")
                        break
                else:
                    print("Failed to fetch data during iteration.")

        except Exception as e:
            print("Error in Recording:", str(e))
    else:
        print("Failed to fetch data from the URL in Recording.")

def Record_Setup():
    print("Starting the Record Setup process")
    global url
    ID_GAME = Getting_Game_Id()
    url = (f"https://ca.1xbet.com/service-api/LiveFeed/GetGameZip?"
           f"id={ID_GAME}&lng=en&isSubGames=true&GroupEvents=true&countevents=250&"
           "grMode=4&topGroups=&country=85&marketType=1")
    print("URL set in Record_Setup:", url)
    response = requests.get(url)
    if response.status_code == 200:
        try:
            print("Fetching initial Time value in Record_Setup...")
            data = response.json()
            Time_val = int(data["Value"]["SC"]["TS"])
            # Here Record_Setup just monitors Time for demonstration.
            for i in range(Time_val):
                response = requests.get(url)
                if response.status_code == 200:
                    data = response.json()
                    Time_val = int(data["Value"]["SC"]["TS"])
                    print("Record_Setup Time value:", Time_val)
                time.sleep(30)
        except Exception as e:
            print(f"Error in Record_Setup: {e}")
    else:
        print("Failed to get data from the server in Record_Setup.")

def delayed_recording():
    """
    This function will run in a separate thread.
    It waits for the initial delay computed from the Time value,
    then starts the Recording process.
    """
    global url
    # Wait until 'url' has been set by Record_Setup
    while not url:
        time.sleep(1)
    # Fetch the initial Time value
    response = requests.get(url)
    if response.status_code == 200:
        data = response.json()
        Time_val = int(data["Value"]["SC"]["TS"])
        # Compute sleep time based on Time_val
        sleep_time = Time_val + 55 if Time_val < 10 else Time_val + 35
        print(f"Delayed recording thread waiting for {sleep_time} seconds before starting Recording.")
        time.sleep(sleep_time)
        Record_Setup()
    else:
        print("Delayed recording thread: Failed to fetch initial data.")

if __name__ == "__main__":
    # Create two threads:
    # 1. Thread to run Record_Setup immediately.
    # 2. Thread to delay (using the computed sleep_time) and then run Recording.
    thread1 = threading.Thread(target=Record_Setup)
    thread2 = threading.Thread(target=delayed_recording)

    thread1.start()
    # A short pause to help ensure that Record_Setup sets the global 'url'
    time.sleep(2)
    thread2.start()

    thread1.join()
    thread2.join()
