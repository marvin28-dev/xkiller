

import time
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException
import openpyxl
from openpyxl import Workbook




workbook = Workbook()
def init_driver():
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_argument("--start-maximized")
    chrome_options.add_argument("--headless")
    driver = webdriver.Chrome(options=chrome_options)
    return driver

# Opening the website
def open_website(driver, url):
    driver.get(url)
    driver.find_element(By.XPATH, '//*[@id="pushfree"]/div/div/div/div/div[2]/div[1]/a').click()

# innitialising data record

def find_element_text(driver, by, value):
    try:
        element = driver.find_element(by, value)
        return element.text
    except NoSuchElementException as e:
        # print(f"Error in finding element with {by}: {value}")
        return None

def next_round(driver):
    time.sleep(100)
    driver.execute_script("document.body.style.zoom = '100%'")
    # Checking for match availability
    for i in range(100):
        print('Checking for match availability for next round...')
        try:
            timings_2 = [
            ('//*[@id="games_content"]/div/div[1]/div/div/div/div[4]/div/div[1]/div/div[2]/div[1]/div[2]/span[2]','//*[@id="games_content"]/div/div[1]/div/div/div/div[4]/div/div[1]/div/div[1]/a'),
            ('//*[@id="games_content"]/div/div[1]/div/div/div/div[5]/div/div[1]/div/div[2]/div[1]/div[2]/span[2]','//*[@id="games_content"]/div/div[1]/div/div/div/div[5]/div/div[1]/div/div[1]/a'),
            ('//*[@id="games_content"]/div/div[1]/div/div/div/div[2]/div/div[1]/div/div[2]/div[1]/div[2]/span[2]','//*[@id="games_content"]/div/div[1]/div/div/div/div[2]/div/div[1]/div/div[1]/a'),
            ('//*[@id="games_content"]/div/div[1]/div/div/div/div[3]/div/div[1]/div/div[2]/div[1]/div[2]/span[2]','//*[@id="games_content"]/div/div[1]/div/div/div/div[3]/div/div[1]/div/div[1]/a'),
            ('//*[@id="games_content"]/div/div[1]/div/div/div/div[6]/div/div[1]/div/div[2]/div[1]/div[2]/span[2]','//*[@id="games_content"]/div/div[1]/div/div/div/div[6]/div/div[1]/div/div[1]/a'),
            ]
            
            for timing_xpath_2, value_2 in timings_2:
                timing_element_2 = driver.find_element(By.XPATH, timing_xpath_2).text
                if ':' in timing_element_2:
                    try:
                        result_time_2 = convert_time(timing_element_2)
                        if result_time_2 < 300:
                            recording(value_2)
                            
                            
                    except ValueError:
                        print("invalid value")
                else:
                    print("no available for recording")
    
            
        except NoSuchElementException as e:
            print("5th match not available yet")    
        
        time.sleep(10)

def record_data(driver):
    current_url_record = driver.current_url
    id_record = match_id(current_url_record)
    try:
        element_time = WebDriverWait(driver, 1000).until(
            EC.presence_of_element_located((By.XPATH, f'//*[@id="{id_record}"]/div/div[1]/div/div[2]/div/div/div[1]/div[4]/span'))
        )
    except:
        print('Error getting the time')
        return

    if element_time:
        time_match = driver.find_element(By.XPATH, f'//*[@id="{id_record}"]/div/div[1]/div/div[2]/div/div/div[1]/div[4]/span').text
        match_time = convert_time(time_match)
        time_left = 600 - match_time
        interval = 30
        num_iterations = time_left // interval
        team1 = driver.find_element(By.XPATH, '//*[@id="allBetsTable"]/div[1]/div[1]/div/div[2]/div[1]/span[1]').text
        team2 = driver.find_element(By.XPATH, '//*[@id="allBetsTable"]/div[1]/div[1]/div/div[2]/div[3]/span[1]').text       

        xpath_value_dict = {
            team1: {
                'TeamXPath': '//*[@id="allBetsTable"]/div[1]/div[1]/div/div[2]/div[1]/span[1]',
                'ValueXPath': '//*[@id="allBetsTable"]/div[1]/div[1]/div/div[2]/div[1]/span[2]/i'
            },
            'Draw': {
                'TeamXPath': '//*[@id="allBetsTable"]/div[1]/div[1]/div/div[2]/div[2]/span[1]',
                'ValueXPath': '//*[@id="allBetsTable"]/div[1]/div[1]/div/div[2]/div[2]/span[2]/i'
            },
            team2: {
                'TeamXPath': '//*[@id="allBetsTable"]/div[1]/div[1]/div/div[2]/div[3]/span[1]',
                'ValueXPath': '//*[@id="allBetsTable"]/div[1]/div[1]/div/div[2]/div[3]/span[2]/i'
            },
            f'{team1} Score':{
                'TeamXPath': '//*[@id="allBetsTable"]/div[1]/div[1]/div/div[2]/div[1]/span[1]',
                'ValueXPath': f'//*[@id="{id_record}"]/div/div[1]/div/div[2]/div/div/div[3]/div[1]/table/tbody/tr[1]/td[2]'
            },
             f'{team2} Score':{
                'TeamXPath': '//*[@id="allBetsTable"]/div[1]/div[1]/div/div[2]/div[1]/span[1]',
                'ValueXPath': f'//*[@id="{id_record}"]/div/div[1]/div/div[2]/div/div/div[3]/div[1]/table/tbody/tr[3]/td[2]'
            },
             'Time':{
                'TeamXPath': '//*[@id="allBetsTable"]/div[1]/div[1]/div/div[2]/div[1]/span[1]',
                'ValueXPath': f'//*[@id="{id_record}"]/div/div[1]/div/div[2]/div/div/div[1]/div[4]'
            }
       
        }
        header = ['Iteration', team1, 'Draw', team2,f'{team1} Score',f'{team2} Score','Time']

        # Load the workbook and select the active sheet
        try:
            wb = openpyxl.load_workbook('recorded_data.xlsx')
            sheet = wb.active
        except FileNotFoundError:
            wb = Workbook()
            sheet = wb.active
        new_sheet_name = f'{id_record}'
        if new_sheet_name in wb.sheetnames:
            sheet = wb[new_sheet_name]
        else:
            sheet = wb.create_sheet(title=new_sheet_name)
        sheet.append(header)
   
        for i in range(num_iterations):
            time.sleep(7)
            print(f'{i} loop has started')
            values_dict = {}
            # Iterate over the xpath_value_dict
            for team, xpaths in xpath_value_dict.items():
                try:
                    # Get the text of the elements using the XPaths
                    team_name = driver.find_element(By.XPATH, xpaths['TeamXPath']).text
                    value = driver.find_element(By.XPATH, xpaths['ValueXPath']).text
                    # Store the values in the values_dict
                    values_dict[team] = value
                except Exception as e:
                    print(f'Error getting data for {team}: {e}')
          
            # Create a new row with the iteration number and default values
            row = [i] + [''] * (len(header) - 1)
            
            for team, value in values_dict.items():
                if team in header:
                    # Find the index of the team in the header to place the value in the correct column
                    idx = header.index(team)
                    row[idx] = value

            # Append the row to the sheet
            sheet.append(row)
            
            # Save the workbook
            wb.save("recorded_data.xlsx")
            
            time.sleep(23)
    next_round(driver)
    

        
# Recording the Data
def recording(link, driver):
    time.sleep(5)
    print('Starting the recording phase')
    try:
        driver.find_element(By.XPATH, link).click()
        driver.execute_script("document.body.style.zoom = '70%'")
        current_url = driver.current_url
        id = match_id(current_url)
        path_first_half = f'//*[@id="{id}"]/div/div[1]/div/div[2]/div/div/div[1]/a'
        locator = (By.XPATH, path_first_half)
        expected_text = '1 HALF'
        try:
            element = WebDriverWait(driver, 1000).until(
            EC.text_to_be_present_in_element(locator, expected_text))
        except:
            print('fatal error waiting for 1st half')
        if element:
            print("your recording has started successfully")
            record_data(driver)
    except Exception as e:
        print(f'little glitch:{e}')

# Getting the match id
def match_id(match_url):
    value_url = match_url.split('/')[-1]
    split_url = value_url.split("-")
    extracted_id = split_url[0]
    return extracted_id

# Converting time to seconds
def convert_time(extracted_time):
    minutes, seconds = map(int, extracted_time.split(':'))
    time_seconds = minutes * 60 + seconds 
    return time_seconds


# Checking for match availability
def check_availability(driver):
    for i in range(100):
        print('Checking for match availability...')
        try:
            timings = [                                                                                            
            ('//*[@id="games_content"]/div/div[1]/div/div/div/div[4]/div/div[1]/div/div[2]/div[1]/div[2]/span[2]','//*[@id="games_content"]/div/div[1]/div/div/div/div[4]/div/div[1]/div/div[1]/a'),
            ('//*[@id="games_content"]/div/div[1]/div/div/div/div[5]/div/div[1]/div/div[2]/div[1]/div[2]/span[2]','//*[@id="games_content"]/div/div[1]/div/div/div/div[5]/div/div[1]/div/div[1]/a'),
            ('//*[@id="games_content"]/div/div[1]/div/div/div/div[2]/div/div[1]/div/div[2]/div[1]/div[2]/span[2]','//*[@id="games_content"]/div/div[1]/div/div/div/div[2]/div/div[1]/div/div[1]/a'),
            ('//*[@id="games_content"]/div/div[1]/div/div/div/div[3]/div/div[1]/div/div[2]/div[1]/div[2]/span[2]','//*[@id="games_content"]/div/div[1]/div/div/div/div[3]/div/div[1]/div/div[1]/a'),
            ('//*[@id="games_content"]/div/div[1]/div/div/div/div[6]/div/div[1]/div/div[2]/div[1]/div[2]/span[2]','//*[@id="games_content"]/div/div[1]/div/div/div/div[6]/div/div[1]/div/div[1]/a'),
            ]
            
            for timing_xpath, value in timings:
                timing_element = driver.find_element(By.XPATH, timing_xpath).text
                if ':' in timing_element:
                    try:
                        result_time = convert_time(timing_element)
                        if result_time < 300:
                            recording(value, driver)
                            
                            
                    except ValueError:
                        print("invalid value")
                else:
                    print("no available for recording")
    
            
        except NoSuchElementException as e:
            print("5th match not available yet")    

        time.sleep(10)

def main():
    driver = init_driver()
    url = "https://1xbet.com/en/live/fifa/2665392-fc-24-3x3-international-masters-league"
    try:
        open_website(driver, url)
        check_availability(driver)
    finally:
        driver.quit()

if __name__ == "__main__":
    main()

